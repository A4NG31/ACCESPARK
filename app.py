import streamlit as st
import pandas as pd
import os
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import io
import base64

# ========================================
# CONFIGURACIÓN DE PÁGINA
# ========================================

st.set_page_config(
    page_title="Validador de Cobros ACCESSPARK",
    page_icon="🅿️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ========================================
# CSS PERSONALIZADO
# ========================================

st.markdown("""
<style>
    .main-header {
        text-align: center;
        color: #2E86AB;
        font-size: 2.5rem;
        font-weight: bold;
        margin-bottom: 2rem;
        padding: 1rem;
        background: linear-gradient(90deg, #f0f8ff, #e6f3ff);
        border-radius: 10px;
        border: 2px solid #2E86AB;
    }
    
    .sub-header {
        color: #A23B72;
        font-size: 1.3rem;
        font-weight: bold;
        margin: 1rem 0;
        padding: 0.5rem;
        background-color: #faf0f5;
        border-radius: 5px;
        border-left: 4px solid #A23B72;
    }
    
    .info-box {
        background-color: #f0f8ff;
        padding: 1rem;
        border-radius: 10px;
        border: 1px solid #d0e0ff;
        margin: 1rem 0;
    }
    
    .success-box {
        background-color: #f0fff0;
        padding: 1rem;
        border-radius: 10px;
        border: 1px solid #90ee90;
        margin: 1rem 0;
    }
    
    .warning-box {
        background-color: #fff8dc;
        padding: 1rem;
        border-radius: 10px;
        border: 1px solid #ffd700;
        margin: 1rem 0;
    }
    
    .metric-container {
        background: white;
        padding: 1rem;
        border-radius: 10px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        border: 1px solid #e0e0e0;
    }
    
    .upload-section {
        background: #fafafa;
        padding: 2rem;
        border-radius: 15px;
        border: 2px dashed #cccccc;
        margin: 1rem 0;
        text-align: center;
    }
    
    .process-button {
        background: linear-gradient(45deg, #2E86AB, #A23B72);
        color: white;
        padding: 0.75rem 2rem;
        border: none;
        border-radius: 25px;
        font-size: 1.1rem;
        font-weight: bold;
        cursor: pointer;
        transition: all 0.3s ease;
    }
    
    .download-button {
        background: linear-gradient(45deg, #28a745, #20c997);
        color: white;
        padding: 0.75rem 2rem;
        border: none;
        border-radius: 25px;
        font-size: 1.1rem;
        font-weight: bold;
        cursor: pointer;
        transition: all 0.3s ease;
        text-decoration: none;
        display: inline-block;
    }
    
    .footer {
        text-align: center;
        padding: 2rem;
        color: #666;
        background-color: #f8f9fa;
        border-radius: 10px;
        margin-top: 2rem;
    }
</style>
""", unsafe_allow_html=True)

st.markdown("""
    <style>
    /* ===== Sidebar ===== */
    [data-testid="stSidebar"] {
        background-color: #1E1E2F !important;
        color: white !important;
        width: 300px !important;
        padding: 20px 10px 20px 10px !important;
        border-right: 1px solid #333 !important;
    }

    [data-testid="stSidebar"] * {
        color: white !important;
    }

    [data-testid="stSidebarNav"] button {
        background: #2E2E3E !important;
        color: white !important;
        border-radius: 6px !important;
    }

    [data-testid="stSidebar"] h1, 
    [data-testid="stSidebar"] h2, 
    [data-testid="stSidebar"] h3 {
        color: #00CFFF !important;
    }
    </style>
""", unsafe_allow_html=True)


# Logo de GoPass con contenedor estilizado
st.markdown("""
<div class="logo-container">
    <img src="https://i.imgur.com/z9xt46F.jpeg"
         style="width: 60%; border-radius: 10px; display: block; margin: 0 auto;" 
         alt="Logo Gopass">
</div>
""", unsafe_allow_html=True)


# ========================================
# FUNCIONES DE PROCESAMIENTO
# ========================================

def procesar_fecha_hora_accesspark(fecha_hora_str):
    """
    Procesa la columna check_in de ACCESSPARK
    Input: '2025-02-27 14:23:00.000'
    Output: fecha='27/02/2025', hora='14:23'
    """
    try:
        if pd.isna(fecha_hora_str):
            return None, None
        
        # Convertir a datetime
        dt = pd.to_datetime(fecha_hora_str, errors='coerce')
        if pd.isna(dt):
            return None, None
        
        # NORMALIZAR al formato DD/MM/YYYY para que coincida con GOPASS
        fecha = dt.strftime('%d/%m/%Y')
        hora = dt.strftime('%H:%M')
        
        return fecha, hora
    except:
        return None, None

def procesar_fecha_hora_gopass(fecha_hora_str):
    """
    Procesa la columna Fecha de entrada de GOPASS
    Input: '28/10/2025  2:57:50 p. m.'
    Output: fecha='28/10/2025', hora='14:57'
    """
    try:
        if pd.isna(fecha_hora_str):
            return None, None
        
        fecha_hora_str = str(fecha_hora_str).strip()
        
        # Intentar parsear con varios formatos
        formatos = [
            '%d/%m/%Y %I:%M:%S %p',  # 28/10/2025 2:57:50 p. m.
            '%d/%m/%Y %H:%M:%S',      # 28/10/2025 14:57:50
            '%d/%m/%Y %I:%M %p',      # 28/10/2025 2:57 p. m.
            '%d/%m/%Y %H:%M',         # 28/10/2025 14:57
        ]
        
        # Limpiar formato de AM/PM en español
        fecha_hora_str = fecha_hora_str.replace(' a. m.', ' AM').replace(' p. m.', ' PM')
        
        dt = None
        for formato in formatos:
            try:
                dt = pd.to_datetime(fecha_hora_str, format=formato, errors='coerce')
                if not pd.isna(dt):
                    break
            except:
                continue
        
        if dt is None or pd.isna(dt):
            # Último intento con parseo automático
            dt = pd.to_datetime(fecha_hora_str, errors='coerce')
        
        if pd.isna(dt):
            return None, None
        
        fecha = dt.strftime('%d/%m/%Y')
        hora = dt.strftime('%H:%M')
        
        return fecha, hora
    except:
        return None, None

def crear_llave(placa, fecha, hora):
    """Crea una llave única combinando placa, fecha y hora"""
    if pd.isna(placa) or pd.isna(fecha) or pd.isna(hora):
        return None
    
    placa_limpia = str(placa).strip().upper().replace(' ', '')
    fecha_limpia = str(fecha).strip()
    hora_limpia = str(hora).strip()
    
    return f"{placa_limpia}|{fecha_limpia}|{hora_limpia}"

def generar_llaves_con_tolerancia(placa, fecha, hora, minutos_tolerancia=10):
    """
    Genera múltiples llaves con tolerancia de tiempo
    Retorna una lista de llaves: [llave_exacta, llave_-10min, llave_-9min, ..., llave_+9min, llave_+10min]
    """
    if pd.isna(placa) or pd.isna(fecha) or pd.isna(hora):
        return []
    
    try:
        # Convertir hora a datetime para hacer operaciones
        hora_base = datetime.strptime(hora, '%H:%M')
        llaves = []
        
        # Generar llaves con tolerancia de -10 a +10 minutos
        for offset in range(-minutos_tolerancia, minutos_tolerancia + 1):
            from datetime import timedelta
            nueva_hora = hora_base + timedelta(minutes=offset)
            hora_str = nueva_hora.strftime('%H:%M')
            llave = crear_llave(placa, fecha, hora_str)
            if llave:
                llaves.append(llave)
        
        return llaves
    except:
        # Si hay error, retornar solo la llave exacta
        llave = crear_llave(placa, fecha, hora)
        return [llave] if llave else []

def leer_archivo(archivo):
    """Lee un archivo Excel o CSV"""
    try:
        nombre = archivo.name.lower()
        if nombre.endswith('.csv'):
            # Leer el contenido como bytes primero
            contenido = archivo.read()
            archivo.seek(0)
            
            # Intentar con diferentes encodings y separadores
            encodings = ['utf-8', 'latin-1', 'iso-8859-1', 'cp1252', 'utf-8-sig']
            separadores = [',', ';', '\t', '|']
            df = None
            
            for encoding in encodings:
                for sep in separadores:
                    try:
                        from io import StringIO
                        texto = contenido.decode(encoding)
                        df_temp = pd.read_csv(StringIO(texto), sep=sep, engine='python')
                        
                        # Verificar si la lectura fue exitosa (más de 1 columna)
                        if len(df_temp.columns) > 1:
                            df = df_temp
                            st.success(f"✅ Archivo CSV leído correctamente con separador '{sep}' y encoding '{encoding}'")
                            break
                    except:
                        continue
                
                if df is not None:
                    break
            
            if df is None:
                # Último intento con detección automática
                archivo.seek(0)
                df = pd.read_csv(archivo, sep=None, engine='python')
            
            # Limpiar nombres de columnas
            df.columns = df.columns.str.strip()
            return df
        else:
            df = pd.read_excel(archivo)
            df.columns = df.columns.str.strip()
            return df
    except Exception as e:
        st.error(f"Error al leer el archivo {archivo.name}: {str(e)}")
        import traceback
        st.error(traceback.format_exc())
        return None

def procesar_archivos_accesspark(archivos_accesspark, archivo_gopass):
    """Procesa los archivos de ACCESSPARK y GOPASS"""
    
    # Leer y concatenar archivos de ACCESSPARK
    dfs_accesspark = []
    for archivo in archivos_accesspark:
        df = leer_archivo(archivo)
        if df is not None:
            dfs_accesspark.append(df)
    
    if not dfs_accesspark:
        st.error("No se pudo leer ningún archivo de ACCESSPARK")
        return None, None
    
    df_accesspark = pd.concat(dfs_accesspark, ignore_index=True)
    
    # Leer archivo de GOPASS
    df_gopass = leer_archivo(archivo_gopass)
    if df_gopass is None:
        return None, None
    
    # Verificar columnas necesarias en ACCESSPARK
    columnas_accesspark = df_accesspark.columns.tolist()
    st.info(f"📋 Columnas encontradas en ACCESSPARK: {columnas_accesspark}")
    
    if 'check_in' not in df_accesspark.columns or 'plate_in' not in df_accesspark.columns:
        st.error(f"❌ El archivo de ACCESSPARK debe contener las columnas 'check_in' y 'plate_in'")
        st.error(f"Columnas actuales: {', '.join(columnas_accesspark)}")
        return None, None
    
    # Verificar columnas necesarias en GOPASS
    columnas_gopass = df_gopass.columns.tolist()
    st.info(f"📋 Columnas encontradas en GOPASS: {columnas_gopass}")
    
    if 'Fecha de entrada' not in df_gopass.columns or 'Placa Vehiculo' not in df_gopass.columns:
        st.error(f"❌ El archivo de GOPASS debe contener las columnas 'Fecha de entrada' y 'Placa Vehiculo'")
        st.error(f"Columnas actuales: {', '.join(columnas_gopass)}")
        return None, None
    
    # Procesar ACCESSPARK
    st.info("📊 Procesando archivos de ACCESSPARK...")
    df_accesspark[['fecha_entrada', 'hora_entrada']] = df_accesspark['check_in'].apply(
        lambda x: pd.Series(procesar_fecha_hora_accesspark(x))
    )
    df_accesspark['llave_exacta'] = df_accesspark.apply(
        lambda row: crear_llave(row['plate_in'], row['fecha_entrada'], row['hora_entrada']), 
        axis=1
    )
    # Generar llaves con tolerancia
    df_accesspark['llaves_tolerancia'] = df_accesspark.apply(
        lambda row: generar_llaves_con_tolerancia(row['plate_in'], row['fecha_entrada'], row['hora_entrada'], 10),
        axis=1
    )
    
    # Procesar GOPASS
    st.info("📊 Procesando archivo de GOPASS...")
    df_gopass[['fecha_entrada', 'hora_entrada']] = df_gopass['Fecha de entrada'].apply(
        lambda x: pd.Series(procesar_fecha_hora_gopass(x))
    )
    df_gopass['llave_exacta'] = df_gopass.apply(
        lambda row: crear_llave(row['Placa Vehiculo'], row['fecha_entrada'], row['hora_entrada']), 
        axis=1
    )
    # Generar llaves con tolerancia
    df_gopass['llaves_tolerancia'] = df_gopass.apply(
        lambda row: generar_llaves_con_tolerancia(row['Placa Vehiculo'], row['fecha_entrada'], row['hora_entrada'], 10),
        axis=1
    )
    
    # Crear conjuntos de llaves para búsqueda rápida (con tolerancia)
    llaves_accesspark = set()
    for llaves_list in df_accesspark['llaves_tolerancia'].dropna():
        llaves_accesspark.update(llaves_list)
    
    llaves_gopass = set()
    for llaves_list in df_gopass['llaves_tolerancia'].dropna():
        llaves_gopass.update(llaves_list)
    
    # Agregar columna de coincidencias en ACCESSPARK (verificar si alguna llave con tolerancia coincide)
    def verificar_coincidencia_accesspark(llaves_list):
        if not llaves_list:
            return 'Llave NO encontrada en GOPASS'
        for llave in llaves_list:
            if llave in llaves_gopass:
                return 'Llave encontrada en GOPASS'
        return 'Llave NO encontrada en GOPASS'
    
    df_accesspark['Estado_Validacion'] = df_accesspark['llaves_tolerancia'].apply(verificar_coincidencia_accesspark)
    
    # Agregar columna de coincidencias en GOPASS (verificar si alguna llave con tolerancia coincide)
    def verificar_coincidencia_gopass(llaves_list):
        if not llaves_list:
            return 'Llave NO encontrada en ACCESSPARK'
        for llave in llaves_list:
            if llave in llaves_accesspark:
                return 'Llave encontrada en ACCESSPARK'
        return 'Llave NO encontrada en ACCESSPARK'
    
    df_gopass['Estado_Validacion'] = df_gopass['llaves_tolerancia'].apply(verificar_coincidencia_gopass)
    
    # Eliminar columnas temporales de llaves con tolerancia antes de exportar
    df_accesspark_export = df_accesspark.drop(columns=['llaves_tolerancia'])
    df_gopass_export = df_gopass.drop(columns=['llaves_tolerancia'])
    
    return df_accesspark_export, df_gopass_export

def crear_excel_resultado(df_accesspark, df_gopass):
    """Crea el archivo Excel con las dos hojas procesadas"""
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_accesspark.to_excel(writer, sheet_name="ACCESSPARK_Procesado", index=False)
        df_gopass.to_excel(writer, sheet_name="GOPASS_Procesado", index=False)
    
    # Aplicar formato condicional
    output.seek(0)
    wb = load_workbook(output)
    
    # Colores para el formato condicional
    verde_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    rojo_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    def aplicar_formato_validacion(ws, nombre_columna):
        """Aplica color verde a encontradas y rojo a no encontradas"""
        col_idx = None
        for cell in ws[1]:
            if cell.value == nombre_columna:
                col_idx = cell.column
                break
        
        if not col_idx:
            return
        
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            if cell.value and 'encontrada en' in str(cell.value) and 'NO' not in str(cell.value):
                cell.fill = verde_fill
            elif cell.value and 'NO encontrada' in str(cell.value):
                cell.fill = rojo_fill
    
    # Aplicar formato a ambas hojas
    aplicar_formato_validacion(wb["ACCESSPARK_Procesado"], "Estado_Validacion")
    aplicar_formato_validacion(wb["GOPASS_Procesado"], "Estado_Validacion")
    
    output_final = io.BytesIO()
    wb.save(output_final)
    output_final.seek(0)
    
    return output_final.getvalue()

# ========================================
# INTERFAZ PRINCIPAL
# ========================================

def main():
    # Header principal
    st.markdown('<div class="main-header">🅿️ Validador de Cobros ACCESSPARK</div>', unsafe_allow_html=True)
    
    # Sidebar con información
    with st.sidebar:
        st.markdown("""
            <div style="
                text-align: center; 
                background: linear-gradient(135deg, #f0f8ff, #e6f3ff);
                padding: 1rem; 
                border-radius: 12px; 
                margin-bottom: 1.5rem;
                box-shadow: 0 2px 6px rgba(0,0,0,0.1);
            ">
                <h2 style="color: #2E86AB; margin: 0;">🅿️ ACCESSPARK</h2>
                <p style="color: #666; margin: 0.5rem 0 0 0;">Sistema de Validación</p>
            </div>
        """, unsafe_allow_html=True)

        st.markdown("### ℹ️ Información del Sistema")
        st.info("Esta aplicación valida y reconcilia automáticamente los registros de ACCESSPARK con la base de datos de GOPASS.")

        st.markdown("### 📋 Funcionalidades")
        st.write("✅ Carga múltiple de archivos")
        st.write("✅ Procesamiento de fechas y horas")
        st.write("✅ Creación de llaves únicas")
        st.write("✅ Validación de coincidencias")
        st.write("✅ Reportes con formato condicional")

        st.markdown("---")
        
        st.markdown("### 📝 Formato de Archivos")
        st.write("**ACCESSPARK:**")
        st.write("- Columnas: check_in, plate_in")
        st.write("- Formato fecha: YYYY-MM-DD HH:MM:SS")
        st.write("- Tolerancia: ±10 minutos")
        
        st.write("**GOPASS:**")
        st.write("- Columnas: Fecha de entrada, Placa Vehiculo")
        st.write("- Formato fecha: DD/MM/YYYY HH:MM:SS")
        st.write("- Tolerancia: ±10 minutos")

    # Sección de carga de archivos
    st.markdown('<div class="sub-header">📤 Carga de Archivos</div>', unsafe_allow_html=True)
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("### 📊 Base ACCESSPARK")
        archivos_accesspark = st.file_uploader(
            "Selecciona uno o varios archivos de ACCESSPARK",
            type=['xlsx', 'xls', 'csv'],
            accept_multiple_files=True,
            key="accesspark"
        )
        if archivos_accesspark:
            st.success(f"✅ {len(archivos_accesspark)} archivo(s) cargado(s)")
            for archivo in archivos_accesspark:
                st.write(f"• {archivo.name}")
    
    with col2:
        st.markdown("### 📊 Base GOPASS")
        archivo_gopass = st.file_uploader(
            "Selecciona el archivo de GOPASS",
            type=['xlsx', 'xls', 'csv'],
            key="gopass"
        )
        if archivo_gopass:
            st.success(f"✅ Archivo cargado: {archivo_gopass.name}")
    
    # Botón de procesamiento
    if archivos_accesspark and archivo_gopass:
        st.markdown("---")
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            if st.button("🚀 VALIDAR COBROS", type="primary", use_container_width=True):
                process_files(archivos_accesspark, archivo_gopass)
    else:
        st.markdown('<div class="warning-box">', unsafe_allow_html=True)
        st.warning("⚠️ Por favor, carga los archivos de ACCESSPARK y GOPASS para continuar con la validación.")
        st.markdown('</div>', unsafe_allow_html=True)

def process_files(archivos_accesspark, archivo_gopass):
    """Maneja el procesamiento de archivos con indicadores de progreso"""
    
    # Barra de progreso
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    try:
        # Paso 1: Procesar archivos
        status_text.text("📊 Procesando archivos...")
        progress_bar.progress(20)
        
        df_accesspark, df_gopass = procesar_archivos_accesspark(archivos_accesspark, archivo_gopass)
        
        if df_accesspark is None or df_gopass is None:
            progress_bar.progress(0)
            status_text.text("❌ Error en el procesamiento")
            return
        
        progress_bar.progress(60)
        status_text.text("🔍 Validando coincidencias...")
        
        # Paso 2: Mostrar estadísticas
        progress_bar.progress(80)
        status_text.text("📈 Generando estadísticas...")
        
        mostrar_estadisticas(df_accesspark, df_gopass)
        
        # Paso 3: Crear archivo de descarga
        progress_bar.progress(90)
        status_text.text("📁 Preparando archivo de descarga...")
        
        excel_data = crear_excel_resultado(df_accesspark, df_gopass)
        
        progress_bar.progress(100)
        status_text.text("✅ ¡Validación completada!")
        
        # Botón de descarga
        st.markdown("---")
        st.markdown('<div class="sub-header">💾 Descargar Resultados</div>', unsafe_allow_html=True)
        
        fecha_actual = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_archivo = f"validacion_accesspark_{fecha_actual}.xlsx"
        
        st.download_button(
            label="📥 DESCARGAR ARCHIVO VALIDADO",
            data=excel_data,
            file_name=nombre_archivo,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True
        )
        
        st.success("🎉 ¡Archivo listo para descargar!")
        
    except Exception as e:
        st.error(f"❌ Error durante el procesamiento: {str(e)}")
        progress_bar.progress(0)
        status_text.text("❌ Error en el procesamiento")

def mostrar_estadisticas(df_accesspark, df_gopass):
    """Muestra estadísticas del procesamiento"""
    
    st.markdown('<div class="sub-header">📊 Estadísticas de Validación</div>', unsafe_allow_html=True)
    
    # Métricas ACCESSPARK
    st.markdown("### 🅿️ Resultados ACCESSPARK")
    col1, col2, col3 = st.columns(3)
    
    total_accesspark = len(df_accesspark)
    encontradas_accesspark = len(df_accesspark[df_accesspark['Estado_Validacion'].str.contains('encontrada en GOPASS', case=False, na=False) & 
                                                ~df_accesspark['Estado_Validacion'].str.contains('NO', case=False, na=False)])
    no_encontradas_accesspark = total_accesspark - encontradas_accesspark
    
    with col1:
        st.markdown('<div class="metric-container">', unsafe_allow_html=True)
        st.metric("📊 Total Registros", total_accesspark)
        st.markdown('</div>', unsafe_allow_html=True)
    
    with col2:
        st.markdown('<div class="metric-container">', unsafe_allow_html=True)
        st.metric("✅ Encontradas en GOPASS", encontradas_accesspark)
        porcentaje = (encontradas_accesspark / total_accesspark * 100) if total_accesspark > 0 else 0
        st.write(f"**{porcentaje:.1f}%** del total")
        st.markdown('</div>', unsafe_allow_html=True)
    
    with col3:
        st.markdown('<div class="metric-container">', unsafe_allow_html=True)
        st.metric("❌ NO Encontradas en GOPASS", no_encontradas_accesspark)
        porcentaje = (no_encontradas_accesspark / total_accesspark * 100) if total_accesspark > 0 else 0
        st.write(f"**{porcentaje:.1f}%** del total")
        st.markdown('</div>', unsafe_allow_html=True)
    
    # Métricas GOPASS
    st.markdown("### 🎫 Resultados GOPASS")
    col1, col2, col3 = st.columns(3)
    
    total_gopass = len(df_gopass)
    encontradas_gopass = len(df_gopass[df_gopass['Estado_Validacion'].str.contains('encontrada en ACCESSPARK', case=False, na=False) & 
                                        ~df_gopass['Estado_Validacion'].str.contains('NO', case=False, na=False)])
    no_encontradas_gopass = total_gopass - encontradas_gopass
    
    with col1:
        st.markdown('<div class="metric-container">', unsafe_allow_html=True)
        st.metric("📊 Total Registros", total_gopass)
        st.markdown('</div>', unsafe_allow_html=True)
    
    with col2:
        st.markdown('<div class="metric-container">', unsafe_allow_html=True)
        st.metric("✅ Encontradas en ACCESSPARK", encontradas_gopass)
        porcentaje = (encontradas_gopass / total_gopass * 100) if total_gopass > 0 else 0
        st.write(f"**{porcentaje:.1f}%** del total")
        st.markdown('</div>', unsafe_allow_html=True)
    
    with col3:
        st.markdown('<div class="metric-container">', unsafe_allow_html=True)
        st.metric("❌ NO Encontradas en ACCESSPARK", no_encontradas_gopass)
        porcentaje = (no_encontradas_gopass / total_gopass * 100) if total_gopass > 0 else 0
        st.write(f"**{porcentaje:.1f}%** del total")
        st.markdown('</div>', unsafe_allow_html=True)
    
    # Resumen visual
    st.markdown("### 📈 Resumen Visual")
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("#### ACCESSPARK")
        data_accesspark = {
            'Estado': ['Encontradas', 'No Encontradas'],
            'Cantidad': [encontradas_accesspark, no_encontradas_accesspark]
        }
        st.bar_chart(pd.DataFrame(data_accesspark).set_index('Estado'))
    
    with col2:
        st.markdown("#### GOPASS")
        data_gopass = {
            'Estado': ['Encontradas', 'No Encontradas'],
            'Cantidad': [encontradas_gopass, no_encontradas_gopass]
        }
        st.bar_chart(pd.DataFrame(data_gopass).set_index('Estado'))

# ========================================
# EJECUTAR APLICACIÓN
# ========================================

if __name__ == "__main__":
    main()
    
    # Footer
    st.markdown("---")
    st.markdown('<div class="footer">💻 Desarrollado por Angel Torres | 🅿️ Validador de Cobros ACCESSPARK | 🚀 Powered by Streamlit</div>', unsafe_allow_html=True)
